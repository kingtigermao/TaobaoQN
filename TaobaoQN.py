# -*- coding: utf-8 -*-
import time
import os
import pyautogui
import pyperclip
from os import remove
from selenium import webdriver
from time import sleep
import tkinter as tk
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from tkinter import filedialog, messagebox

class Qn:
    @staticmethod
    def capture_screen_shot(screenshot_name, path, x=1450, y=160, width=300, height=20):
        time.sleep(5)
        # 截取屏幕上的一部分图像（例如，左上角为 (x, y)，宽度为 width，高度为 height）
        # x, y, width, height = 1200, 240, 400, 100    #公司电脑
        # x, y, width, height = 1320, 200, 400, 80  # 自己电脑
        #x, y, width, height = 1450, 150, 400, 100  # 台式机
        screenshot = pyautogui.screenshot(region=(int(x), int(y), int(width), int(height)))
        new_x = x + 1
        new_y = y + 1
        pyautogui.doubleClick(x=new_x, y=new_y, clicks=2)
        pyautogui.hotkey('ctrl', 'c')
        copied_value = pyperclip.paste()

        # 保存截图为文件
        sc_path = f"{path}\\{screenshot_name}.png"
        screenshot.save(sc_path)
        return sc_path, copied_value

    @staticmethod
    def get_qn_list(url, usename, start_time, end_time, path, x, y, width, height, t_type):
        # Open and Login in
        print("代码运行中，请千万不要动鼠标，不要操作浏览器，将持续几分钟，谢谢！")
        print("执行完毕后将自动关闭浏览器页面")
        sleep(1)
        driver = webdriver.Firefox()
        try:
            driver.get(url)
            driver.maximize_window()
            # Wait Until Page Contains Element
            driver.implicitly_wait(180)
            if t_type == "Tianmao":
                driver.find_element(By.XPATH, "//span[text() = '待发货订单数']")
            else:
                driver.find_element(By.XPATH, "//div[@title='待付款']")
            driver.find_element(By.XPATH,
                                "//*[starts-with(@class, 'FirstClassMenu--navItemText') and text() = '交易']").click()
            time.sleep(3)
            # 点击展开所有筛选项
            driver.find_element(By.XPATH,"//*[starts-with(@class, 'search-form_foldable-cursor') and text() = '展开所有筛选项']").click()
            time.sleep(2)
            # 选择创建时间起始时间+确认
            driver.find_element(By.XPATH,
                                "//label[text() = '创建时间']").click()
            time.sleep(1)
            driver.find_element(By.XPATH, "(//input[@placeholder = 'YYYY-MM-DD'])[1]").send_keys(start_time)
            time.sleep(1)
            driver.find_element(By.XPATH, "(//input[@placeholder = 'YYYY-MM-DD'])[2]").send_keys(end_time)
            time.sleep(1)
            # 点确认
            driver.find_element(By.XPATH, "//span[text() = '确定']").click()
            time.sleep(1)
            #点搜索订单
            driver.find_element(By.XPATH, "//span[text() = '搜索订单']").click()
            time.sleep(5)
            total_count = driver.find_element(By.XPATH,
                                              "/html/body/div[1]/div[3]/div[2]/div/div/div/div/div/div[5]/div/span").text
            print(f"total_count:{total_count}")
            timestamp = time.time()
            time_struct = time.localtime(timestamp)
            formatted_time = time.strftime("%Y%m%d%H%M%S", time_struct)
            if not os.path.exists(path):
                os.mkdir(path)
            file_name = f"{path}\\order_info_{t_type}_{usename}_{start_time}_{end_time}_{formatted_time}.xlsx"
            if int(total_count) < 15:
                total_page = 1
            elif int(total_count)%15 == 0:
                total_page = int(int(total_count) / 15)
            else:
                total_page = int(int(total_count) / 15) + 1
            page = total_page
            add_title_info_into_excel(file_name)
            first_time_flag = True
            for i in range(1, total_page + 1):
                page -= 1
                print(f"page:{page}")
                count = len(driver.find_elements(By.XPATH, "//div[@class='next-table-body']/table")) + 1
                # Loop through elements on the page
                for index in range(1, count):
                    order_info = driver.find_element(By.XPATH,
                                                     f"//div[@class='next-table-body']/table[{index}]/tbody/tr[1]/td/div/div/div[1]").text
                    order_time = str(order_info).split("\n")[1].split("：")[1]
                    order_id = str(order_info).split("\n")[0].split("：")[1]
                    ww_name = str(order_info).split("\n")[2]
                    # 商品名称
                    order_name = driver.find_element(By.XPATH, f"//div[@class='next-table-body']/table[{index}]/tbody/tr[2]/td[1]/div/div/div/a").text
                    # 交易状态
                    order_status = driver.find_element(By.XPATH,
                                                     f"//div[@class='next-table-body']/table[{index}]/tbody/tr[2]/td[5]/div/div/div[1]/div[1]").text
                    # 订单金额
                    order_value = driver.find_element(By.XPATH,
                                                     f"//div[@class='next-table-body']/table[{index}]/tbody/tr[2]/td[6]/div/div/div[1]").text
                    order_value = str(order_value).split("￥")[1]
                    #点击旺旺头像
                    driver.find_element(By.XPATH, f"//div[@class='next-table-body']/table[{index}]/tbody/tr[1]/td/div/div/div[1]/span[2]/span/span[1]/a").click()
                    #等待手动点击允许
                    if first_time_flag == True:
                        sleep(3)
                    else:
                        sleep(1)
                    image_file_path, copied_value = Qn.capture_screen_shot(f"{i}_{count}_{index}_{formatted_time}", path, x, y, width, height)
                    add_qianniu_into_to_txt(
                        order_time, order_id, ww_name, order_name, order_status, order_value,
                        file_name, image_file_path, index, i, copied_value)
                if page == 0:
                    break
                driver.find_element(By.XPATH, "//span[text() = '下一页']").click()
                sleep(5)
            print(f"您获取的信息已经写在此路径下：{file_name}")
        finally:
            print("finish")
            driver.close()

def add_title_info_into_excel( file_name):
    # add order info
    try:
        # 尝试加载已有的工作簿
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        workbook = Workbook()
    # 选择活动的工作表
    sheet = workbook.active
    data_to_insert = ["日期", "订单号", "旺旺名", "金额", "预留", "预留", "预留", "预留", "旺旺名2", "商品名称", "订单状态"]
    # 在第N行插入数据
    sheet.append(data_to_insert)
    # 保存工作簿到文件
    workbook.save(file_name)

def add_qianniu_into_to_txt(
        order_time, order_id, ww_name, order_name, order_status, order_value, file_name, image_file_path, index, i, copied_value):
    # add order info
    try:
        # 尝试加载已有的工作簿
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        workbook = Workbook()
    # 选择活动的工作表
    sheet = workbook.active
    kong = ""
    data_to_insert = [order_time, order_id, ww_name, order_value, kong, kong, kong, kong, copied_value, order_name, order_status]
    # 在第N行插入数据
    sheet.append(data_to_insert)
    # 保存工作簿到文件
    workbook.save(file_name)

    # add img
    line = (i - 1) * 15 + index + 1
    image_column = 'H'
    wb = load_workbook(file_name)  # 打开excel工作簿
    ws = wb.active  # 获取活跃工作表
    try:
        img = Image(image_file_path)  # 获取图片
        img.width, img.height = (300, 20)  # 设置图片大小
        # 调整表格列宽和行高
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions[image_column].width = 40
        ws.row_dimensions[line].height = 25
        ws.add_image(img, anchor=image_column + str(line))  # 插入对应单元格
    except Exception as e:
        print(e)
    wb.save(file_name)  # 保存
    remove(image_file_path)
    print('save..')


url = "https://loginmyseller.taobao.com/?from=taobaoindex&f=top&style=&sub=true&redirect_url=https%3A%2F%2Fqn.taobao.com%2Fhome.htm%2Ftrade-platform%2Ftp%2Fsold"
# start_time = "2023-11-18"
# end_time = "2023-11-19"
# path = "C:\\Testttt"
# usename = "sa"

#Qn.get_qn_list(url, usename, start_time, end_time, path)
def on_button_click():
    if not entry_username.get() or not entry_start_time.get() or not entry_end_time.get() or not entry_storage_path.get()\
            or not entry_x.get() or not entry_y.get() or not entry_height.get() or not entry_width.get() \
            or not entry_c_type.get():
        result_label.config(text="请填写所有必要信息")
    else:
        messagebox.showinfo("注意！", "代码即将运行，请点击OK后千万不要动鼠标，不要操作浏览器，将持续几分钟，谢谢！")
        result_label.config(
            Qn.get_qn_list(url, entry_username.get(), entry_start_time.get(), entry_end_time.get(),
                           entry_storage_path.get(), entry_x.get(), entry_y.get(), entry_width.get(),
                           entry_height.get(), entry_c_type.get()))
        # Show a message box after processing
        messagebox.showinfo("操作完成", "您获取的信息已经写在以下路径下:\n" + entry_storage_path.get())

def browse_folder():
    folder_path = filedialog.askdirectory()
    entry_storage_path.delete(0, tk.END)  # Clear previous content
    entry_storage_path.insert(tk.END, folder_path)

# 创建主窗口
root = tk.Tk()
root.title("自动获取淘宝/天猫店订单信息程序")
root.geometry("1000x600")  # 设置默认窗口大小, 宽*高

# 添加用户名标签和文本框
label_username = tk.Label(root, text="店铺名")
label_username.grid(row=0, column=0, padx=10, pady=10)
entry_username = tk.Entry(root)
entry_username.grid(row=0, column=1, padx=10, pady=10)
entry_username.insert(tk.END, "Huinuo")

# 添加店铺类型标签和文本框
label_c_type = tk.Label(root, text="店铺类型：Tianmao 或者 Taobao")
label_c_type.grid(row=1, column=0, padx=10, pady=10)
entry_c_type = tk.Entry(root)
entry_c_type.grid(row=1, column=1, padx=10, pady=10)
entry_c_type.insert(tk.END, "Taobao")

# 添加开始时间标签和文本框
label_start_time = tk.Label(root, text="开始时间")
label_start_time.grid(row=2, column=0, padx=10, pady=10)
entry_start_time = tk.Entry(root)
entry_start_time.grid(row=2, column=1, padx=10, pady=10)
entry_start_time.insert(tk.END, "2023-11-18")

# 添加结束时间标签和文本框
label_end_time = tk.Label(root, text="结束时间")
label_end_time.grid(row=3, column=0, padx=10, pady=10)
entry_end_time = tk.Entry(root)
entry_end_time.grid(row=3, column=1, padx=10, pady=10)
entry_end_time.insert(tk.END, "2023-11-19")

# 添加x值标签和文本框
label_x = tk.Label(root, text="截图x轴值")
label_x.grid(row=4, column=0, padx=10, pady=10)
entry_x = tk.Entry(root)
entry_x.grid(row=4, column=1, padx=10, pady=10)
entry_x.insert(tk.END, "1450")

# 添加y值标签和文本框
label_y = tk.Label(root, text="截图y轴值")
label_y.grid(row=4, column=2, padx=10, pady=10)
entry_y = tk.Entry(root)
entry_y.grid(row=4, column=3, padx=10, pady=10)
entry_y.insert(tk.END, "160")

# 添加width值标签和文本框
label_width = tk.Label(root, text="截图宽度")
label_width.grid(row=5, column=0, padx=10, pady=10)
entry_width = tk.Entry(root)
entry_width.grid(row=5, column=1, padx=10, pady=10)
entry_width.insert(tk.END, "300")

# 添加height值标签和文本框
label_height = tk.Label(root, text="截图高度")
label_height.grid(row=5, column=2, padx=10, pady=10)
entry_height = tk.Entry(root)
entry_height.grid(row=5, column=3, padx=10, pady=10)
entry_height.insert(tk.END, "20")

# 添加存放路径标签、文本框和浏览按钮
label_storage_path = tk.Label(root, text="存放路径")
label_storage_path.grid(row=6, column=0, padx=10, pady=10)
entry_storage_path = tk.Entry(root)
entry_storage_path.grid(row=6, column=1, padx=10, pady=10)
button_browse = tk.Button(root, text="浏览", command=browse_folder)
button_browse.grid(row=6, column=2, padx=10, pady=10)

# 添加按钮，并绑定点击事件
button = tk.Button(root, text="生成", command=on_button_click)
button.grid(row=7, column=0, columnspan=2, pady=10)  # columnspan用于横跨两列

# 添加用于显示结果的标签
result_label = tk.Label(root, text="")
result_label.grid(row=8, column=0, columnspan=3, pady=10)

# 启动主循环
root.mainloop()